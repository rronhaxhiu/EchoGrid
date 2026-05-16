"""
Serialization utilities for converting between domain objects and
storable/exportable representations (dicts, JSON).
"""

from io import BytesIO
import json
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..domain.event import SimEvent
from ..domain.influence import InfluenceMatrix
from ..domain.run import SimulationRun, Snapshot
from ..domain.tile import Tile
from ..domain.world import World
from ..application.world_initializer import WorldInitializer


class RunSerializer:
    """Converts SimulationRun ↔ plain dicts / JSON strings."""

    @staticmethod
    def reconstruct_run(model: Any, events: List[SimEvent]) -> SimulationRun:
        """
        Rebuild a SimulationRun from a DB model + event list.
        Uses stored current_world_state for fast resumption (no replay needed).
        """
        current_tick = model.current_tick
        # Re-queue any event whose tick is strictly in the future so it fires
        # correctly when this reconstructed engine runs its next tick.  Without
        # this, Postgres-backed services (new instance per request) would load
        # an empty pending_events list and silently drop user-injected events.
        pending = [e for e in events if e.tick > current_tick]

        run = SimulationRun(
            id=model.id,
            seed=model.seed,
            variables=list(model.variables),
            global_initial_values=dict(model.global_initial_values),
            hex_radius=model.hex_radius,
            spatial_decay=float(model.spatial_decay),
            influence_config=dict(model.influence_config),
            current_tick=current_tick,
            event_log=list(events),
            pending_events=pending,
            created_at=model.created_at.isoformat() if model.created_at else None,
        )

        influence = InfluenceMatrix.from_dict(model.influence_config)
        run.influence = influence

        if model.current_world_state:
            world = _world_from_snapshot(model.current_world_state)
        else:
            world = WorldInitializer.create(
                seed=model.seed,
                hex_radius=model.hex_radius,
                variables=list(model.variables),
                global_initial_values=dict(model.global_initial_values),
            )

        run.world = world
        return run

    @staticmethod
    def export_run(run: SimulationRun) -> Dict[str, Any]:
        """Export a full run to a JSON-serializable dict."""
        return {
            "meta": run.meta_dict(),
            "world_state": run.world.snapshot(),
            "global_state": run.world.get_global_state(),
            "event_log": [e.to_dict() for e in run.event_log],
            "snapshots": [s.to_dict() for s in run.snapshots],
        }

    @staticmethod
    def export_run_json(run: SimulationRun) -> str:
        return json.dumps(RunSerializer.export_run(run), indent=2)

    @staticmethod
    def export_run_xlsx(run: SimulationRun) -> bytes:
        """Export a full run to an Excel workbook."""
        return build_run_export_workbook(RunSerializer.export_run(run))


def build_run_export_workbook(export_data: Dict[str, Any]) -> bytes:
    """Build an XLSX workbook from the JSON-serializable run export shape."""
    wb = Workbook()

    meta = export_data.get("meta", {})
    world_state = export_data.get("world_state", {})
    global_state = export_data.get("global_state", {})
    event_log = export_data.get("event_log", [])
    snapshots = export_data.get("snapshots", [])
    if not isinstance(meta, dict):
        meta = {}
    if not isinstance(world_state, dict):
        world_state = {}
    if not isinstance(global_state, dict):
        global_state = {}
    if not isinstance(event_log, list):
        event_log = []
    if not isinstance(snapshots, list):
        snapshots = []

    metadata_ws = wb.active
    metadata_ws.title = "Metadata"
    _append_rows(
        metadata_ws,
        ["field", "value"],
        [
            ["run id", meta.get("id")],
            ["seed", meta.get("seed")],
            ["hex radius", meta.get("hex_radius")],
            ["current tick", meta.get("current_tick")],
            ["tile count", meta.get("tile_count", len(world_state))],
            ["event count", meta.get("event_count", len(event_log))],
            ["snapshot count", meta.get("snapshot_count", len(snapshots))],
            ["created_at", meta.get("created_at")],
            ["variables", _json_cell(meta.get("variables", []))],
            ["spatial_decay", meta.get("spatial_decay")],
        ],
    )

    global_ws = wb.create_sheet("Global State")
    _append_rows(
        global_ws,
        ["variable", "value"],
        [[variable, value] for variable, value in sorted(global_state.items())],
    )

    world_ws = wb.create_sheet("World State")
    variable_names = _world_variable_names(world_state, meta.get("variables", []))
    world_rows = []
    for key, variables in sorted(world_state.items(), key=lambda item: _tile_sort_key(item[0])):
        if not isinstance(variables, dict):
            variables = {}
        q, r = _parse_tile_key(key)
        world_rows.append([q, r, *[variables.get(name) for name in variable_names]])
    _append_rows(world_ws, ["q", "r", *variable_names], world_rows)

    events_ws = wb.create_sheet("Events")
    _append_rows(
        events_ws,
        ["id", "tick", "name", "source", "delta_map", "target_tiles"],
        [
            [
                event.get("id"),
                event.get("tick"),
                event.get("name"),
                event.get("source"),
                _json_cell(event.get("delta_map", {})),
                _json_cell(event.get("target_tiles", [])),
            ]
            for event in event_log
            if isinstance(event, dict)
        ],
    )

    snapshots_ws = wb.create_sheet("Snapshots")
    _append_rows(
        snapshots_ws,
        ["id", "run_id", "tick", "is_diff", "created_at", "state"],
        [
            [
                snapshot.get("id"),
                snapshot.get("run_id"),
                snapshot.get("tick"),
                snapshot.get("is_diff"),
                snapshot.get("created_at"),
                _json_cell(snapshot.get("state", {})),
            ]
            for snapshot in snapshots
            if isinstance(snapshot, dict)
        ],
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        _autosize_columns(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _append_rows(ws: Worksheet, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def _json_cell(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _world_variable_names(world_state: Dict[str, Any], preferred_order: List[str]) -> List[str]:
    names = list(preferred_order) if isinstance(preferred_order, list) else []
    seen = set(names)
    for variables in world_state.values():
        if not isinstance(variables, dict):
            continue
        for name in variables:
            if name not in seen:
                names.append(name)
                seen.add(name)
    return names


def _parse_tile_key(key: str) -> tuple[int | None, int | None]:
    try:
        q_str, r_str = key.split(",", 1)
        return int(q_str), int(r_str)
    except (AttributeError, ValueError):
        return None, None


def _tile_sort_key(key: str) -> tuple[int, int, str]:
    q, r = _parse_tile_key(key)
    if q is None or r is None:
        return (0, 0, str(key))
    return (q, r, str(key))


def _autosize_columns(ws: Worksheet) -> None:
    for column in ws.columns:
        width = 10
        column_letter = column[0].column_letter
        for cell in column:
            value = cell.value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, 60))
        ws.column_dimensions[column_letter].width = width


def _world_from_snapshot(snapshot: Dict[str, Any]) -> World:
    """Reconstruct a World object from a snapshot dict."""
    tiles: Dict[tuple, Tile] = {}
    for key, variables in snapshot.items():
        q_str, r_str = key.split(",")
        q, r = int(q_str), int(r_str)
        tiles[(q, r)] = Tile(q=q, r=r, variables=dict(variables))

    world = World(tiles=tiles)
    world.build_neighbor_cache()
    return world
